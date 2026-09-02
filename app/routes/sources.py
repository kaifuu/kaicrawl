# -*- coding: utf-8 -*-
"""数据源管理：列表 / 新增 / 编辑 / 启停 / 删除 / 从Excel重导入 / 立即抓取 / 导出Excel。"""
import io
import re
from datetime import date

from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, current_app, send_file)

from sqlalchemy import or_

from ..extensions import db
from ..models import Source, Article, Task, CrawlLog, CrawlLogLine
from .articles import remove_article_files, prune_empty_dirs
from ..sources import PARSER_REGISTRY
from .. import scheduler_jobs, excel_sync
from config import EXCEL_PATH, MAX_ARTICLES_BACKFILL

bp = Blueprint("sources", __name__)

PER_PAGE_OPTIONS = [10, 20, 50, 100]


def _parser_options():
    """[(key, 标签)]，供表单下拉。"""
    return [(key, getattr(cls, "site_name", "") or key)
            for key, cls in PARSER_REGISTRY.items()]


def _since_from_form():
    """读取表单 since（起始日期），返回 (since_date_or_None, 错误消息_or_None)。

    合法：空（=不限日期）或 YYYY-MM-DD；非法给出错误消息。
    """
    since = (request.form.get("since") or "").strip()
    if since and not re.match(r"^\d{4}-\d{2}-\d{2}$", since):
        return None, "起始日期格式应为 YYYY-MM-DD"
    return since or None, None


def _limit_from_form():
    """读取表单 limit（本次最多抓取篇数），返回 int 或 None（None=用解析器默认）。

    空 → None；非法或 <1 → None（容错，不报错打断）；超过上限则截到 MAX_ARTICLES_BACKFILL。
    """
    raw = (request.form.get("limit") or "").strip()
    if not raw:
        return None
    try:
        n = int(raw)
    except ValueError:
        return None
    if n < 1:
        return None
    return min(n, MAX_ARTICLES_BACKFILL)


@bp.route("/")
def index():
    q = (request.args.get("q") or "").strip()
    cat = (request.args.get("category") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    if per_page not in PER_PAGE_OPTIONS:
        per_page = 20
    query = Source.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Source.name.like(like), Source.url.like(like)))
    if cat:
        query = query.filter_by(category=cat)
    pagination = query.order_by(Source.category, Source.id).paginate(
        page=page, per_page=per_page, error_out=False)
    categories = sorted({c for (c,) in
                         Source.query.with_entities(Source.category).distinct() if c})
    return render_template("sources.html", sources=pagination.items,
                           pagination=pagination, q=q, category=cat,
                           categories=categories, per_page=per_page,
                           per_page_options=PER_PAGE_OPTIONS)


@bp.route("/export")
def export():
    """导出数据源列表为 Excel：跟随当前搜索/分类筛选，无筛选时导出全部。

    列名与导入 Excel（新闻/栏目/来源/列表区域XPath/分页URL模板/渲染模式/备注）
    保持一致，重合列可直接粘回导入表；另含界面上可配的其余全部字段。
    """
    q = (request.args.get("q") or "").strip()
    cat = (request.args.get("category") or "").strip()
    query = Source.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Source.name.like(like), Source.url.like(like)))
    if cat:
        query = query.filter_by(category=cat)
    sources = query.order_by(Source.category, Source.id).all()

    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据源"
    headers = ["新闻", "栏目", "来源", "来源类型", "解析器", "作者策略",
               "列表区域XPath", "列表项日期XPath", "时间来源行XPath(旧)",
               "时间XPath", "来源XPath", "作者XPath", "正文区域XPath", "备选正文区域XPath",
               "分页URL模板", "渲染模式", "启用", "备注", "创建时间"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for s in sources:
        ws.append([
            s.category, s.name, s.url, s.source_type, s.parser_key,
            s.author_policy, s.list_xpath, s.date_xpath, s.meta_xpath,
            s.time_xpath, s.source_xpath, s.author_xpath,
            s.content_xpath, s.content_xpath_alt, s.page_url_pattern, s.render_mode,
            "是" if s.enabled else "否", s.remark or "",
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
        ])
    widths = [14, 22, 48, 10, 10, 12, 30, 26, 36, 26, 26, 26, 30, 30, 26, 10, 6, 24, 17]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"数据源导出_{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/new", methods=["GET", "POST"])
def new():
    if request.method == "POST":
        s = Source(
            category=request.form.get("category", "").strip(),
            name=request.form.get("name", "").strip(),
            url=request.form.get("url", "").strip(),
            source_type=request.form.get("source_type", "website").strip(),
            parser_key=request.form.get("parser_key", "bjdch").strip(),
            author_policy=request.form.get("author_policy", "").strip(),
            content_xpath=request.form.get("content_xpath", "").strip(),
            content_xpath_alt=request.form.get("content_xpath_alt", "").strip(),
            list_xpath=request.form.get("list_xpath", "").strip(),
            date_xpath=request.form.get("date_xpath", "").strip(),
            time_xpath=request.form.get("time_xpath", "").strip(),
            source_xpath=request.form.get("source_xpath", "").strip(),
            author_xpath=request.form.get("author_xpath", "").strip(),
            page_url_pattern=request.form.get("page_url_pattern", "").strip(),
            render_mode=request.form.get("render_mode", "static").strip() or "static",
            remark=request.form.get("remark", "").strip(),
            enabled=request.form.get("enabled") == "on",
        )
        if not s.category or not s.name:
            flash("分类与名称必填", "danger")
            return render_template("source_form.html", source=s,
                                   parser_options=_parser_options())
        db.session.add(s)
        db.session.commit()
        flash("数据源已添加", "success")
        return redirect(url_for("sources.index"))
    return render_template("source_form.html", source=None,
                           parser_options=_parser_options())


@bp.route("/<int:sid>/edit", methods=["GET", "POST"])
def edit(sid):
    s = db.session.get(Source, sid)
    if not s:
        flash("数据源不存在", "danger")
        return redirect(url_for("sources.index"))
    if request.method == "POST":
        s.category = request.form.get("category", "").strip()
        s.name = request.form.get("name", "").strip()
        s.url = request.form.get("url", "").strip()
        s.source_type = request.form.get("source_type", "website").strip()
        s.parser_key = request.form.get("parser_key", "bjdch").strip()
        s.author_policy = request.form.get("author_policy", "").strip()
        s.content_xpath = request.form.get("content_xpath", "").strip()
        s.content_xpath_alt = request.form.get("content_xpath_alt", "").strip()
        s.list_xpath = request.form.get("list_xpath", "").strip()
        s.date_xpath = request.form.get("date_xpath", "").strip()
        # meta_xpath 已从表单移除（被时间/来源/作者三区域取代），编辑时保留旧值作兜底
        s.time_xpath = request.form.get("time_xpath", "").strip()
        s.source_xpath = request.form.get("source_xpath", "").strip()
        s.author_xpath = request.form.get("author_xpath", "").strip()
        s.page_url_pattern = request.form.get("page_url_pattern", "").strip()
        s.render_mode = request.form.get("render_mode", "static").strip() or "static"
        s.remark = request.form.get("remark", "").strip()
        s.enabled = request.form.get("enabled") == "on"
        db.session.commit()
        flash("已保存", "success")
        return redirect(url_for("sources.index"))
    return render_template("source_form.html", source=s,
                           parser_options=_parser_options())


@bp.route("/<int:sid>/toggle", methods=["POST"])
def toggle(sid):
    s = db.session.get(Source, sid)
    if s:
        s.enabled = not s.enabled
        db.session.commit()
        flash(f"已{'启用' if s.enabled else '禁用'}：{s.name}", "success")
    return redirect(url_for("sources.index"))


@bp.route("/<int:sid>/delete", methods=["POST"])
def delete(sid):
    """删除数据源，并级联清理其文章（含 WORD/图片文件）、任务与运行日志。

    articles/tasks.source_id 非空，靠关系默认级联会把 source_id 置 NULL 而
    撞 NOT NULL 约束（500），必须显式先删子记录；子记录在 session 标记删除后，
    父记录删除不再尝试置空外键。
    """
    s = db.session.get(Source, sid)
    if not s:
        flash("数据源不存在", "danger")
        return redirect(url_for("sources.index"))
    if CrawlLog.query.filter_by(source_id=s.id, status="running").first():
        flash("该数据源正在抓取，请先停止任务再删除", "warning")
        return redirect(url_for("sources.index"))

    articles = s.articles.all()
    n_files, img_dirs = 0, set()
    for a in articles:
        n, dirs = remove_article_files(a)
        n_files += n
        img_dirs |= dirs
        db.session.delete(a)

    log_ids = [r[0] for r in db.session.query(CrawlLog.id)
               .filter_by(source_id=s.id).all()]
    if log_ids:
        CrawlLogLine.query.filter(CrawlLogLine.log_id.in_(log_ids)).delete(synchronize_session=False)
        CrawlLog.query.filter(CrawlLog.id.in_(log_ids)).delete(synchronize_session=False)

    tasks = list(s.tasks)
    for t in tasks:
        scheduler_jobs.remove_task(t.id)
        db.session.delete(t)

    db.session.delete(s)
    db.session.commit()
    prune_empty_dirs(img_dirs)
    flash(f"已删除「{s.name}」及其 {len(articles)} 篇文章、{n_files} 个文件、"
          f"{len(tasks)} 个任务、{len(log_ids)} 条运行日志", "success")
    return redirect(url_for("sources.index"))


@bp.route("/<int:sid>/run", methods=["POST"])
def run(sid):
    """立即抓取某来源（后台线程）。可选起始日期 since（YYYY-MM-DD）回溯抓取。"""
    s = db.session.get(Source, sid)
    if not s:
        flash("数据源不存在", "danger")
        return redirect(url_for("sources.index"))
    since, err = _since_from_form()
    if err:
        flash(err, "danger")
        return redirect(url_for("sources.index"))
    limit = _limit_from_form()
    scheduler_jobs.run_now(current_app._get_current_object(), s.id,
                           since_date=since, limit=limit)
    flash(f"已触发抓取：{s.name}（后台执行中，请稍后查看日志）", "info")
    return redirect(url_for("sources.index"))


@bp.route("/<int:sid>/run-overwrite", methods=["POST"])
def run_overwrite(sid):
    """立即覆盖抓取某来源：对已存在的文章删旧记录+旧 WORD 后重新抓取生成。

    可选起始日期 since（YYYY-MM-DD）回溯抓取。
    """
    s = db.session.get(Source, sid)
    if not s:
        flash("数据源不存在", "danger")
        return redirect(url_for("sources.index"))
    since, err = _since_from_form()
    if err:
        flash(err, "danger")
        return redirect(url_for("sources.index"))
    limit = _limit_from_form()
    scheduler_jobs.run_now(current_app._get_current_object(), s.id,
                           overwrite=True, since_date=since, limit=limit)
    flash(f"已触发覆盖抓取：{s.name}（后台执行中，请稍后查看日志）", "info")
    return redirect(url_for("sources.index"))


@bp.route("/run-all", methods=["POST"])
def run_all():
    """一键抓取：对所有启用数据源按既定配置触发后台抓取。

    弹窗里的 since（YYYY-MM-DD 回溯）/ limit（每源最多篇数）对所有启用来源统一应用；
    禁用来源按其「既定配置」跳过。每个来源各起一个后台线程并发执行。
    """
    since, err = _since_from_form()
    if err:
        flash(err, "danger")
        return redirect(url_for("sources.index"))
    limit = _limit_from_form()
    app = current_app._get_current_object()
    enabled = Source.query.filter_by(enabled=True).order_by(
        Source.category, Source.id).all()
    if not enabled:
        flash("没有可抓取的启用数据源", "warning")
        return redirect(url_for("sources.index"))
    for s in enabled:
        scheduler_jobs.run_now(app, s.id, since_date=since, limit=limit)
    flash(f"已触发 {len(enabled)} 个数据源抓取（后台执行中，请稍后查看日志）", "info")
    return redirect(url_for("sources.index"))


@bp.route("/reimport", methods=["POST"])
def reimport():
    n = excel_sync.import_from_excel(EXCEL_PATH, replace=False)
    flash(f"已从 Excel 同步数据源（共 {n} 条）", "success")
    return redirect(url_for("sources.index"))
